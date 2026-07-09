from __future__ import annotations

import datetime as dt
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATA_DIR = Path("Data")
OUTPUT_PREFIX = "Baocao_TH"

REQUIRED_FIELDS = ("Ngày", "Thời gian", "Trưởng ca", "Nội dung")


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def normalize_header(value) -> str:
    text = strip_accents(str(value or "")).lower().strip()
    text = re.sub(r"\s+", " ", text)
    if text in {"stt", "so thu tu"}:
        return "STT"
    if text in {"ngay"}:
        return "Ngày"
    if text in {"gio", "thoi gian", "thoi gian/gio"}:
        return "Thời gian"
    if text in {"truong ca"}:
        return "Trưởng ca"
    if text in {"noi dung"}:
        return "Nội dung"
    return str(value or "").strip()


def normalize_date(value) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).split(" ")[0].strip() if value else ""


def normalize_time(value) -> str:
    if isinstance(value, dt.datetime):
        return value.time().strftime("%H:%M:%S")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)):
        seconds = round(float(value) * 24 * 60 * 60)
        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours % 24:02d}:{minutes:02d}:{seconds:02d}"
    return str(value).strip() if value else ""


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def workshop_name(path: Path) -> str:
    match = re.search(r"(\d+)", path.stem)
    return f"Phân xưởng {match.group(1)}" if match else path.stem


def output_path() -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(f"{OUTPUT_PREFIX}_{timestamp}.xlsx")


def make_report_content(shift_leader, content) -> str:
    leader = str(shift_leader).strip() if not is_blank(shift_leader) else "[Thiếu trưởng ca]"
    report = str(content).strip() if not is_blank(content) else "[Thiếu nội dung]"
    return f"{leader}: {report}"


def source_files() -> list[Path]:
    return sorted(DATA_DIR.glob("*.xlsx"))


def read_source_rows() -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    rows: list[dict] = []
    warnings: list[dict] = []
    file_summaries: list[dict] = []
    column_summaries: list[dict] = []
    seen_columns: set[tuple[str, str]] = set()

    for path in source_files():
        workbook = openpyxl.load_workbook(path, data_only=True)
        worksheet = workbook.active
        px_name = workshop_name(path)

        raw_headers = [cell.value for cell in worksheet[1]]
        normalized_headers = [normalize_header(header) for header in raw_headers]
        header_index = {header: index for index, header in enumerate(normalized_headers) if header}

        file_summaries.append(
            {
                "file": path.name,
                "workshop": px_name,
                "sheet": worksheet.title,
                "rows": max(worksheet.max_row - 1, 0),
                "columns": worksheet.max_column,
            }
        )

        for raw_header, normalized_header in zip(raw_headers, normalized_headers):
            key = (str(raw_header or ""), normalized_header)
            if key not in seen_columns:
                seen_columns.add(key)
                column_summaries.append(
                    {
                        "original": str(raw_header or ""),
                        "normalized": normalized_header,
                        "meaning": column_meaning(normalized_header),
                    }
                )

        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            value_map = {
                header: values[index] if index < len(values) else None
                for header, index in header_index.items()
            }
            row = {
                "source_file": path.name,
                "source_row": excel_row,
                "source_stt": value_map.get("STT"),
                "date": normalize_date(value_map.get("Ngày")),
                "time": normalize_time(value_map.get("Thời gian")),
                "workshop": px_name,
                "shift_leader": value_map.get("Trưởng ca"),
                "content": value_map.get("Nội dung"),
            }
            rows.append(row)

            required_values = {
                "Ngày": row["date"],
                "Thời gian": row["time"],
                "Trưởng ca": row["shift_leader"],
                "Nội dung": row["content"],
            }
            for field_name, field_value in required_values.items():
                if is_blank(field_value):
                    warnings.append(
                        {
                            "file": path.name,
                            "row": excel_row,
                            "field": field_name,
                            "issue": f"Thiếu {field_name.lower()}",
                        }
                    )

    return rows, warnings, file_summaries, column_summaries


def column_meaning(header: str) -> str:
    meanings = {
        "STT": "Số thứ tự trong file nguồn",
        "Ngày": "Ngày làm việc",
        "Thời gian": "Thời gian bắt đầu ca",
        "Trưởng ca": "Người phụ trách phân xưởng",
        "Nội dung": "Nội dung công việc đạt được",
    }
    return meanings.get(header, "Cột ngoài cấu trúc chuẩn")


def sorted_report_rows(rows: list[dict]) -> list[dict]:
    return sorted(rows, key=lambda row: (row["date"], row["time"], row["workshop"]))


def write_report(
    rows: list[dict],
    warnings: list[dict],
    file_summaries: list[dict],
    column_summaries: list[dict],
    destination: Path,
) -> None:
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "Tong hop"
    report_sheet.append(["STT", "Ngày", "Thời gian", "Tên phân xưởng", "Nội dung"])

    for index, row in enumerate(sorted_report_rows(rows), start=1):
        report_sheet.append(
            [
                index,
                row["date"],
                row["time"],
                row["workshop"],
                make_report_content(row["shift_leader"], row["content"]),
            ]
        )

    source_sheet = workbook.create_sheet("Buoc 1")
    source_sheet.append(["Tên file", "Tên phân xưởng", "Sheet", "Số dòng dữ liệu", "Số cột"])
    for summary in file_summaries:
        source_sheet.append(
            [
                summary["file"],
                summary["workshop"],
                summary["sheet"],
                summary["rows"],
                summary["columns"],
            ]
        )

    column_sheet = workbook.create_sheet("Cau truc cot")
    column_sheet.append(["Tên cột gốc", "Tên cột chuẩn hóa", "Ý nghĩa"])
    for summary in column_summaries:
        column_sheet.append(
            [
                summary["original"],
                summary["normalized"],
                summary["meaning"],
            ]
        )

    warning_sheet = workbook.create_sheet("Canh bao du lieu")
    warning_sheet.append(["File", "Dòng Excel", "Cột", "Vấn đề"])
    for warning in warnings:
        warning_sheet.append(
            [
                warning["file"],
                warning["row"],
                warning["field"],
                warning["issue"],
            ]
        )

    style_workbook(workbook)
    workbook.save(destination)


def style_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")

    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            width = min(max(max_length + 2, 12), 55)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = width

        worksheet.freeze_panes = "A2"


def print_summary(
    rows: list[dict],
    warnings: list[dict],
    file_summaries: list[dict],
    destination: Path,
) -> None:
    dates = {row["date"] for row in rows if row["date"]}
    workshops = {row["workshop"] for row in rows if row["workshop"]}

    print("Tóm tắt bước 1:")
    for summary in file_summaries:
        print(
            f"- {summary['file']}: {summary['rows']} dòng, "
            f"{summary['columns']} cột, sheet {summary['sheet']}"
        )
    print(f"- Cảnh báo dữ liệu thiếu: {len(warnings)}")

    print("Tóm tắt bước 2:")
    print(f"- Số ngày làm việc: {len(dates)}")
    print(f"- Số phân xưởng: {len(workshops)}")
    print(f"- Số dòng sau tổng hợp: {len(rows)}")
    print("- Sắp xếp theo Ngày, Thời gian, Tên phân xưởng")

    print("Tóm tắt bước 3:")
    print(f"- Đã tạo file Excel tổng hợp: {destination}")


def main() -> None:
    rows, warnings, file_summaries, column_summaries = read_source_rows()
    destination = output_path()
    write_report(rows, warnings, file_summaries, column_summaries, destination)
    print_summary(rows, warnings, file_summaries, destination)


if __name__ == "__main__":
    main()
