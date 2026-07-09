from __future__ import annotations

import datetime as dt
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATA_DIR = Path("Data")
OUTPUT_PREFIX = "Baocao_TH"


def bo_dau_tieng_viet(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def chuan_hoa_ten_cot(value) -> str:
    text = bo_dau_tieng_viet(str(value or "")).lower().strip()
    text = re.sub(r"\s+", " ", text)

    if text in {"stt", "so thu tu"}:
        return "STT"
    if text == "ngay":
        return "Ngày"
    if text in {"gio", "thoi gian", "thoi gian/gio"}:
        return "Thời gian"
    if text == "truong ca":
        return "Trưởng ca"
    if text == "noi dung":
        return "Nội dung"
    return str(value or "").strip()


def chuan_hoa_ngay(value) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).split(" ")[0].strip() if value else ""


def chuan_hoa_thoi_gian(value) -> str:
    if isinstance(value, dt.datetime):
        return value.time().strftime("%H:%M:%S")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)):
        seconds = round(float(value) * 24 * 60 * 60)
        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours % 24:02d}:{minutes:02d}:{seconds:02d}"
    return str(value).strip() if value else ""


def thieu_du_lieu(value) -> bool:
    return value is None or str(value).strip() == ""


def ten_phan_xuong(path: Path) -> str:
    match = re.search(r"(\d+)", path.stem)
    return f"Phân xưởng {match.group(1)}" if match else path.stem


def ten_file_dau_ra() -> Path:
    ngay_gio = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(f"{OUTPUT_PREFIX}_{ngay_gio}.xlsx")


def y_nghia_cot(header: str) -> str:
    meanings = {
        "STT": "Số thứ tự trong file nguồn",
        "Ngày": "Ngày làm việc",
        "Thời gian": "Thời gian bắt đầu ca",
        "Trưởng ca": "Người phụ trách phân xưởng",
        "Nội dung": "Nội dung công việc đạt được",
    }
    return meanings.get(header, "Cột ngoài cấu trúc chuẩn")


def noi_dung_bao_cao(truong_ca, noi_dung) -> str:
    leader = str(truong_ca).strip() if not thieu_du_lieu(truong_ca) else "[Thiếu trưởng ca]"
    content = str(noi_dung).strip() if not thieu_du_lieu(noi_dung) else "[Thiếu nội dung]"
    return f"{leader}: {content}"


def danh_sach_file_nguon() -> list[Path]:
    return sorted(DATA_DIR.glob("*.xlsx"))


def buoc_1_doc_du_lieu() -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    du_lieu: list[dict] = []
    canh_bao: list[dict] = []
    tom_tat_file: list[dict] = []
    cau_truc_cot: list[dict] = []
    cot_da_gap: set[tuple[str, str]] = set()

    for path in danh_sach_file_nguon():
        workbook = openpyxl.load_workbook(path, data_only=True)
        worksheet = workbook.active
        phan_xuong = ten_phan_xuong(path)

        cot_goc = [cell.value for cell in worksheet[1]]
        cot_chuan = [chuan_hoa_ten_cot(header) for header in cot_goc]
        vi_tri_cot = {header: index for index, header in enumerate(cot_chuan) if header}

        tom_tat_file.append(
            {
                "file": path.name,
                "phan_xuong": phan_xuong,
                "sheet": worksheet.title,
                "so_dong": max(worksheet.max_row - 1, 0),
                "so_cot": worksheet.max_column,
            }
        )

        for raw_header, normalized_header in zip(cot_goc, cot_chuan):
            key = (str(raw_header or ""), normalized_header)
            if key not in cot_da_gap:
                cot_da_gap.add(key)
                cau_truc_cot.append(
                    {
                        "cot_goc": str(raw_header or ""),
                        "cot_chuan": normalized_header,
                        "y_nghia": y_nghia_cot(normalized_header),
                    }
                )

        for dong_excel, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row_values = {
                header: values[index] if index < len(values) else None
                for header, index in vi_tri_cot.items()
            }
            row = {
                "file": path.name,
                "dong_excel": dong_excel,
                "stt_nguon": row_values.get("STT"),
                "ngay": chuan_hoa_ngay(row_values.get("Ngày")),
                "thoi_gian": chuan_hoa_thoi_gian(row_values.get("Thời gian")),
                "phan_xuong": phan_xuong,
                "truong_ca": row_values.get("Trưởng ca"),
                "noi_dung": row_values.get("Nội dung"),
            }
            du_lieu.append(row)

            cac_truong_can_kiem_tra = {
                "Ngày": row["ngay"],
                "Thời gian": row["thoi_gian"],
                "Trưởng ca": row["truong_ca"],
                "Nội dung": row["noi_dung"],
            }
            for ten_cot, gia_tri in cac_truong_can_kiem_tra.items():
                if thieu_du_lieu(gia_tri):
                    canh_bao.append(
                        {
                            "file": path.name,
                            "dong_excel": dong_excel,
                            "cot": ten_cot,
                            "van_de": f"Thiếu {ten_cot.lower()}",
                        }
                    )

    return du_lieu, canh_bao, tom_tat_file, cau_truc_cot


def buoc_2_tong_hop_theo_ngay(du_lieu: list[dict]) -> list[dict]:
    du_lieu_sap_xep = sorted(
        du_lieu,
        key=lambda row: (row["ngay"], row["thoi_gian"], row["phan_xuong"]),
    )

    bang_tong_hop = []
    for stt, row in enumerate(du_lieu_sap_xep, start=1):
        bang_tong_hop.append(
            {
                "STT": stt,
                "Ngày": row["ngay"],
                "Thời gian": row["thoi_gian"],
                "Tên phân xưởng": row["phan_xuong"],
                "Nội dung": noi_dung_bao_cao(row["truong_ca"], row["noi_dung"]),
            }
        )
    return bang_tong_hop


def buoc_3_tao_excel(
    bang_tong_hop: list[dict],
    canh_bao: list[dict],
    tom_tat_file: list[dict],
    cau_truc_cot: list[dict],
    destination: Path,
) -> None:
    workbook = Workbook()

    sheet_tong_hop = workbook.active
    sheet_tong_hop.title = "Tong hop"
    sheet_tong_hop.append(["STT", "Ngày", "Thời gian", "Tên phân xưởng", "Nội dung"])
    for row in bang_tong_hop:
        sheet_tong_hop.append(
            [
                row["STT"],
                row["Ngày"],
                row["Thời gian"],
                row["Tên phân xưởng"],
                row["Nội dung"],
            ]
        )

    sheet_buoc_1 = workbook.create_sheet("Buoc 1")
    sheet_buoc_1.append(["Tên file", "Tên phân xưởng", "Sheet", "Số dòng dữ liệu", "Số cột"])
    for row in tom_tat_file:
        sheet_buoc_1.append(
            [
                row["file"],
                row["phan_xuong"],
                row["sheet"],
                row["so_dong"],
                row["so_cot"],
            ]
        )

    sheet_cau_truc = workbook.create_sheet("Cau truc cot")
    sheet_cau_truc.append(["Tên cột gốc", "Tên cột chuẩn hóa", "Ý nghĩa"])
    for row in cau_truc_cot:
        sheet_cau_truc.append([row["cot_goc"], row["cot_chuan"], row["y_nghia"]])

    sheet_canh_bao = workbook.create_sheet("Canh bao du lieu")
    sheet_canh_bao.append(["File", "Dòng Excel", "Cột", "Vấn đề"])
    for row in canh_bao:
        sheet_canh_bao.append([row["file"], row["dong_excel"], row["cot"], row["van_de"]])

    dinh_dang_workbook(workbook)
    workbook.save(destination)


def dinh_dang_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")

    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            width = min(max(max_length + 2, 12), 55)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = width

        worksheet.freeze_panes = "A2"


def in_tom_tat(
    du_lieu: list[dict],
    bang_tong_hop: list[dict],
    canh_bao: list[dict],
    tom_tat_file: list[dict],
    destination: Path,
) -> None:
    so_ngay = len({row["ngay"] for row in du_lieu if row["ngay"]})
    so_phan_xuong = len({row["phan_xuong"] for row in du_lieu if row["phan_xuong"]})

    print("Tóm tắt bước 1 - Đọc dữ liệu:")
    for row in tom_tat_file:
        print(
            f"- {row['file']}: {row['so_dong']} dòng, "
            f"{row['so_cot']} cột, sheet {row['sheet']}"
        )
    print(f"- Cảnh báo dữ liệu thiếu: {len(canh_bao)}")

    print("Tóm tắt bước 2 - Tổng hợp theo ngày:")
    print(f"- Số ngày làm việc: {so_ngay}")
    print(f"- Số phân xưởng: {so_phan_xuong}")
    print(f"- Số dòng sau tổng hợp: {len(bang_tong_hop)}")
    print("- Quy tắc sắp xếp: Ngày, Thời gian, Tên phân xưởng")

    print("Tóm tắt bước 3 - Tạo Excel tổng hợp:")
    print(f"- Đã tạo file: {destination}")


def main() -> None:
    du_lieu, canh_bao, tom_tat_file, cau_truc_cot = buoc_1_doc_du_lieu()
    bang_tong_hop = buoc_2_tong_hop_theo_ngay(du_lieu)
    destination = ten_file_dau_ra()
    buoc_3_tao_excel(bang_tong_hop, canh_bao, tom_tat_file, cau_truc_cot, destination)
    in_tom_tat(du_lieu, bang_tong_hop, canh_bao, tom_tat_file, destination)


if __name__ == "__main__":
    main()
