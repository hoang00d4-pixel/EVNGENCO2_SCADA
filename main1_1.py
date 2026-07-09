from __future__ import annotations

import datetime as dt
import logging
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATA_DIR = Path("Data")
LOG_DIR = Path("logs")
OUTPUT_PREFIX = "Baocao_TH"
FONT_NAME = "Times New Roman"
FONT_SIZE = 12
LOGGER = logging.getLogger("tong_hop_bao_cao")


def setup_logging() -> Path:
    LOG_DIR.mkdir(exist_ok=True)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = LOG_DIR / f"{OUTPUT_PREFIX}_{timestamp}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8")],
        force=True,
    )
    return log_file


def remove_vietnamese_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def normalize_column_name(value) -> str:
    text = remove_vietnamese_accents(str(value or "")).lower().strip()
    text = re.sub(r"\s+", " ", text)

    if text in {"stt", "so thu tu"}:
        return "STT"
    if text == "ngay":
        return "Ngày"
    if text in {"gio", "thoi gian", "thoi gian/gio"}:
        return "Thời gian"
    if text == "truong ca":
        return "Trưởng ca"
    if text == "noi dung":
        return "Nội dung"
    return str(value or "").strip()


def normalize_date(value) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).split(" ")[0].strip() if value else ""


def normalize_time(value) -> str:
    if isinstance(value, dt.datetime):
        return value.time().strftime("%H:%M:%S")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)):
        total_seconds = round(float(value) * 24 * 60 * 60)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours % 24:02d}:{minutes:02d}:{seconds:02d}"
    return str(value).strip() if value else ""


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def get_workshop_name(path: Path) -> str:
    match = re.search(r"(\d+)", path.stem)
    return f"Phân xưởng {match.group(1)}" if match else path.stem


def create_output_path() -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(f"{OUTPUT_PREFIX}_{timestamp}.xlsx")


def get_source_files() -> list[Path]:
    return sorted(path for path in DATA_DIR.glob("*.xlsx") if not path.name.startswith("~$"))


def get_column_meaning(column_name: str) -> str:
    meanings = {
        "STT": "Số thứ tự trong file nguồn",
        "Ngày": "Ngày làm việc",
        "Thời gian": "Thời gian bắt đầu ca",
        "Trưởng ca": "Người phụ trách phân xưởng",
        "Nội dung": "Nội dung công việc đạt được",
    }
    return meanings.get(column_name, "Cột ngoài cấu trúc chuẩn")


def build_report_content(shift_leader, content) -> str:
    leader = str(shift_leader).strip() if not is_blank(shift_leader) else "[Thiếu trưởng ca]"
    work_content = str(content).strip() if not is_blank(content) else "[Thiếu nội dung]"
    return f"{leader}: {work_content}"


def step_1_read_data() -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    normalized_rows: list[dict] = []
    warnings: list[dict] = []
    file_summary: list[dict] = []
    column_summary: list[dict] = []
    seen_columns: set[tuple[str, str]] = set()
    files = get_source_files()

    LOGGER.info("Bước 1 - Bắt đầu đọc dữ liệu từ thư mục %s", DATA_DIR)
    LOGGER.info("Số file .xlsx phát hiện: %s", len(files))

    for path in files:
        workbook = openpyxl.load_workbook(path, data_only=True)
        worksheet = workbook.active
        workshop = get_workshop_name(path)
        LOGGER.info("Đọc file %s, sheet %s, phân xưởng %s", path.name, worksheet.title, workshop)

        original_headers = [cell.value for cell in worksheet[1]]
        normalized_headers = [normalize_column_name(header) for header in original_headers]
        column_index = {
            header: index for index, header in enumerate(normalized_headers) if header
        }

        file_summary.append(
            {
                "file": path.name,
                "workshop": workshop,
                "sheet": worksheet.title,
                "row_count": max(worksheet.max_row - 1, 0),
                "column_count": worksheet.max_column,
            }
        )

        for original, normalized in zip(original_headers, normalized_headers):
            key = (str(original or ""), normalized)
            if key not in seen_columns:
                seen_columns.add(key)
                column_summary.append(
                    {
                        "original": str(original or ""),
                        "normalized": normalized,
                        "meaning": get_column_meaning(normalized),
                    }
                )

        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            value_by_column = {
                header: values[index] if index < len(values) else None
                for header, index in column_index.items()
            }
            row = {
                "source_file": path.name,
                "excel_row": excel_row,
                "source_stt": value_by_column.get("STT"),
                "date": normalize_date(value_by_column.get("Ngày")),
                "time": normalize_time(value_by_column.get("Thời gian")),
                "workshop": workshop,
                "shift_leader": value_by_column.get("Trưởng ca"),
                "content": value_by_column.get("Nội dung"),
            }
            normalized_rows.append(row)

            required_values = {
                "Ngày": row["date"],
                "Thời gian": row["time"],
                "Trưởng ca": row["shift_leader"],
                "Nội dung": row["content"],
            }
            for column_name, column_value in required_values.items():
                if is_blank(column_value):
                    LOGGER.warning(
                        "Thiếu dữ liệu: file=%s, dòng=%s, cột=%s",
                        path.name,
                        excel_row,
                        column_name,
                    )
                    warnings.append(
                        {
                            "file": path.name,
                            "excel_row": excel_row,
                            "column": column_name,
                            "issue": f"Thiếu {column_name.lower()}",
                        }
                    )

    LOGGER.info("Bước 1 - Hoàn tất: %s dòng, %s cảnh báo", len(normalized_rows), len(warnings))
    return normalized_rows, warnings, file_summary, column_summary


def step_2_summarize_by_date(normalized_rows: list[dict]) -> list[dict]:
    LOGGER.info("Bước 2 - Bắt đầu tổng hợp theo ngày")
    sorted_rows = sorted(
        normalized_rows,
        key=lambda row: (row["date"], row["time"], row["workshop"]),
    )

    result = []
    for index, row in enumerate(sorted_rows, start=1):
        result.append(
            {
                "STT": index,
                "Ngày": row["date"],
                "Thời gian": row["time"],
                "Tên phân xưởng": row["workshop"],
                "Nội dung": build_report_content(row["shift_leader"], row["content"]),
            }
        )

    LOGGER.info("Bước 2 - Hoàn tất: %s dòng sau tổng hợp", len(result))
    return result


def step_3_create_excel(
    final_rows: list[dict],
    warnings: list[dict],
    file_summary: list[dict],
    column_summary: list[dict],
    output_file: Path,
) -> None:
    LOGGER.info("Bước 3 - Bắt đầu tạo workbook Excel: %s", output_file)
    workbook = Workbook()

    report_sheet = workbook.active
    report_sheet.title = "Tong hop"
    report_sheet.append(["STT", "Ngày", "Thời gian", "Tên phân xưởng", "Nội dung"])
    for row in final_rows:
        report_sheet.append(
            [
                row["STT"],
                row["Ngày"],
                row["Thời gian"],
                row["Tên phân xưởng"],
                row["Nội dung"],
            ]
        )

    step_1_sheet = workbook.create_sheet("Buoc 1")
    step_1_sheet.append(["Tên file", "Tên phân xưởng", "Sheet", "Số dòng dữ liệu", "Số cột"])
    for row in file_summary:
        step_1_sheet.append(
            [
                row["file"],
                row["workshop"],
                row["sheet"],
                row["row_count"],
                row["column_count"],
            ]
        )

    column_sheet = workbook.create_sheet("Cau truc cot")
    column_sheet.append(["Tên cột gốc", "Tên cột chuẩn hóa", "Ý nghĩa"])
    for row in column_summary:
        column_sheet.append([row["original"], row["normalized"], row["meaning"]])

    warning_sheet = workbook.create_sheet("Canh bao du lieu")
    warning_sheet.append(["File", "Dòng Excel", "Cột", "Vấn đề"])
    for row in warnings:
        warning_sheet.append([row["file"], row["excel_row"], row["column"], row["issue"]])

    format_workbook(workbook)
    workbook.save(output_file)
    LOGGER.info("Bước 3 - Đã lưu file Excel: %s", output_file)


def format_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for cell in worksheet[1]:
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            width = min(max(max_length + 2, 12), 55)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = width

        worksheet.freeze_panes = "A2"


def print_summary(
    normalized_rows: list[dict],
    final_rows: list[dict],
    warnings: list[dict],
    file_summary: list[dict],
    output_file: Path,
    log_file: Path,
) -> None:
    work_dates = {row["date"] for row in normalized_rows if row["date"]}
    workshops = {row["workshop"] for row in normalized_rows if row["workshop"]}

    LOGGER.info("Tóm tắt bước 1 - Số file đã đọc: %s", len(file_summary))
    print("Tóm tắt bước 1:")
    print("- Danh sách file đã đọc:")
    for row in file_summary:
        print(
            f"  + {row['file']}: {row['row_count']} dòng, "
            f"{row['column_count']} cột, sheet {row['sheet']}"
        )
    print(f"- Cảnh báo dữ liệu thiếu: {len(warnings)}")

    LOGGER.info(
        "Tóm tắt bước 2 - Số ngày=%s, số phân xưởng=%s, số dòng=%s",
        len(work_dates),
        len(workshops),
        len(final_rows),
    )
    print("Tóm tắt bước 2:")
    print(f"- Số ngày làm việc: {len(work_dates)}")
    print(f"- Số phân xưởng: {len(workshops)}")
    print(f"- Số dòng sau tổng hợp: {len(final_rows)}")
    print("- Quy tắc sắp xếp: Ngày, Thời gian, Tên phân xưởng")

    LOGGER.info("Tóm tắt bước 3 - File đầu ra: %s", output_file)
    print("Tóm tắt bước 3:")
    print(f"- File Excel tổng hợp: {output_file}")
    print(f"- File log: {log_file}")
    print(f"- Font: {FONT_NAME}, cỡ chữ {FONT_SIZE}")


def main() -> None:
    log_file = setup_logging()
    LOGGER.info("Bắt đầu chạy quy trình tổng hợp báo cáo")
    try:
        normalized_rows, warnings, file_summary, column_summary = step_1_read_data()
        final_rows = step_2_summarize_by_date(normalized_rows)
        output_file = create_output_path()
        step_3_create_excel(final_rows, warnings, file_summary, column_summary, output_file)
        print_summary(normalized_rows, final_rows, warnings, file_summary, output_file, log_file)
        LOGGER.info("Hoàn tất quy trình tổng hợp báo cáo")
    except Exception:
        LOGGER.exception("Quy trình tổng hợp báo cáo bị lỗi")
        print(f"- File log lỗi: {log_file}")
        raise


if __name__ == "__main__":
    main()
